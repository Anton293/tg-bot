from sqlalchemy import create_engine, MetaData, Table
import pandas as pd

import datetime
import openpyxl
from openpyxl import Workbook, load_workbook

engine = create_engine('sqlite:///data/database.db')
metadata = MetaData()
metadata.reflect(bind=engine)
vacancies_table = Table('vacancies', metadata, autoload=True, autoload_with=engine)


async def create_xlsx_table() -> None:
    conn = engine.connect()
    select_query = f""" 
        SELECT *
        FROM vacancies
        WHERE datetime >= datetime('now', 'start of day');
    """
    сursor = conn.execute(select_query)
    result = сursor.fetchall()
    conn.close()

    if result == []:
        return

    df = pd.DataFrame(result, columns=result[0].keys())
    df['datetime'] = pd.to_datetime(df['datetime'])
    df['datetime'] = df['datetime'] + pd.DateOffset(hours=3)
    df['datetime'] = df['datetime'].dt.strftime('%Y-%m-%d %H:%M')
    df = df.drop(columns=['id'])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vacancies'
    ws.append(df.columns.tolist())
    for row in df.iterrows():
        ws.append(row[1].tolist())

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10

    for cell in ws['A1:C1']:
        for c in cell:
            c.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    wb.save('data/vacancies_data.xlsx')
